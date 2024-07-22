from django.shortcuts import render
from tienda.models import Producto
from django.views.generic import TemplateView,View
from openpyxl import Workbook
from django.http.response import HttpResponse
from django.utils import timezone
from tienda.utils import render_pdf
# Create your views here.

def tienda (request):
    product = Producto.objects.all()
    return render(request, "tienda.html", {'productos':product})

def reporte (request):
    producto = Producto.objects.all()
    return render(request, "listado.html", {'productos':producto})

class ReporteProductoExel(TemplateView):
    def get(self, request, *args, **kwargs):
        productos = Producto.objects.all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Productos"
        ws['A1'] = "Reporte Productos"
        ws.merge_cells('A1:E1')
        ws['A3'] = 'Id'
        ws['B3'] = 'Producto'
        ws['C3'] = 'Precio'
        ws['D3'] = 'Imagen'
        ws['E3'] = 'Fecha de Creacion'
        cont = 4

        for producto in productos:
            ws.cell(row=cont, column=1).value = producto.id
            ws.cell(row=cont, column=2).value = producto.nombre
            ws.cell(row=cont, column=3).value = producto.precio
            ws.cell(row=cont, column=4).value = producto.imagen.url
            ws.cell(row=cont, column=5).value = timezone.localtime(producto.created).strftime('%d/%m/%Y')
            cont += 1

        nombre_archivo = "ReportesProductosExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


def listado (request):
    producto = Producto.objects.all()
    return render(request, "listado.html", {'productos':producto})

class ListaProductoPdf(View):
    def get(self, request, *args, **kwargs):
        producto = Producto.objects.all()
        data = {
            "productos": producto
        }
        pdf = render_pdf('tienda/listado.html',data)
        return HttpResponse(pdf,content_type='application/pdf')